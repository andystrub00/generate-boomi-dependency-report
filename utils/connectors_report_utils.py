import csv
import requests
import argparse
from datetime import datetime
from utils.utils import chunk_list
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


def get_sample_deployments_data():

    return [
        {
            "@type": "DeployedPackage",
            "deploymentId": "7a65bc0a-92b1-4823-95b0-a5e70bab7fb2",
            "version": 43,
            "packageId": "ee6a2d33-7394-4535-b1ca-ea75cb46b612",
            "packageVersion": "9.2",
            "environmentId": "e994923b-34fe-4e3c-9d6e-5d586bccfe52",
            "componentId": "118ced80-f2e1-47ce-a784-576923cc49f2",
            "componentVersion": 586,
            "componentType": "process",
            "deployedDate": "2025-03-24T14:58:18Z",
            "deployedBy": "andy.strubhar@argano.com",
            "notes": "Updates to containerize reusable subprocesses as well as use specific process prop for records to go downstream.",
            "active": True,
            "branchName": "main",
        },
        {
            "@type": "DeployedPackage",
            "deploymentId": "fb1501ec-a749-47f7-a14d-da1ab344b3f1",
            "version": 32,
            "packageId": "5457806d-381f-4243-9a1b-b25263125850",
            "packageVersion": "13.4.0",
            "environmentId": "e994923b-34fe-4e3c-9d6e-5d586bccfe52",
            "componentId": "162c4dfd-d465-4431-a4a0-7a74d2d6e5ce",
            "componentVersion": 370,
            "componentType": "process",
            "deployedDate": "2025-09-26T17:51:46Z",
            "deployedBy": "mamiller@sc.younglife.org",
            "notes": "This packaged component will not update SF MissionUnits if there is an existing SF MissionUnit that is Closed. It will send an email notification with info about the SupeOrgs that exist as Closed. It WILL update territory records associated with the closed Mission Unit. It will now update Visible in Community based on the existence of an ACTIVE CC in WD",
            "active": True,
            "branchName": "main",
        },
        {
            "@type": "DeployedPackage",
            "deploymentId": "c648c6f4-408d-4f6d-9640-5de3cc420c8e",
            "version": 14,
            "packageId": "12d15c31-7f77-4deb-bc5d-3f32d732f5f5",
            "packageVersion": "6.6",
            "environmentId": "e994923b-34fe-4e3c-9d6e-5d586bccfe52",
            "componentId": "16ea5d09-a34e-4141-b3df-489ea0da3d5b",
            "componentVersion": 159,
            "componentType": "process",
            "deployedDate": "2025-06-02T20:44:39Z",
            "deployedBy": "andy.strubhar@argano.com",
            "notes": 'Added new field to YL API Candidate Change Request payload "JobCode", which is a 1:1 mapping from the Workday Candidate Change payload field "Job_Code"',
            "active": True,
            "branchName": "main",
        },
        {
            "@type": "DeployedPackage",
            "deploymentId": "ddf30609-0f84-49f9-8af5-9706be795f79",
            "version": 1,
            "packageId": "cf205054-e1e2-4e6d-8257-891f4fa9de84",
            "packageVersion": "1.0",
            "environmentId": "e994923b-34fe-4e3c-9d6e-5d586bccfe52",
            "componentId": "dda2b326-9bb5-4e51-b0e1-5673461356db",
            "componentVersion": 4,
            "componentType": "process",
            "deployedDate": "2025-03-26T21:49:07Z",
            "deployedBy": "mamiller@sc.younglife.org",
            "notes": 'Created Process Route Shape for "Write to Snowflake Audit Table" and "Transform and Upload WCX Registrants"',
            "active": True,
            "branchName": "main",
        },
    ]


def parse_connectors_report_command_line_args():
    """
    Parses command-line arguments for connector report processing.

    Arguments:
    -a, --account_name: Optional; Boomi account nickname which will be appended to the .env filename,
                            e.g -a younglife will look for env vars in younglife.env
    -e, --environment: Optional; Name of the environment to filter connectors by.
    Returns:
    argparse.Namespace: Parsed arguments.
    """
    parser = argparse.ArgumentParser(description="Process connector information.")
    parser.add_argument(
        "-a",
        "--account_name",
        type=str,
        default="",
        help="Boomi account name for env file, e.g -a argano will look for argano.env",
    )
    parser.add_argument(
        "-e",
        "--environment",
        type=str,
        help="Name of the environment to filter connectors by.",
    )

    args = parser.parse_args()

    return args


class BoomiMetadataExtractor:
    """
    Extracts integration metadata from Boomi AtomSphere API
    """

    def __init__(self, account_id: str, username: str, password: str):
        """
        Initialize the Boomi API client

        Args:
            account_id: Your Boomi account ID
            username: API username (format: username@account_id)
            password: API password/token
        """
        self.account_id = account_id
        self.base_url = f"https://api.boomi.com/api/rest/v1/{account_id}"

        self.auth = (f"BOOMI_TOKEN.{username}", password)

        self.headers = {
            "Content-Type": "application/json",
            "Accept": "application/json",
        }

    def query_deployed_packages(
        self, environment_name: str = None
    ) -> list[dict[str, any]]:
        """
        Query all deployed packages (processes) in the account

        Args:
            environment_name: Optional environment name to filter (e.g., "AWS Molecule Prod")

        Returns:
            List of deployed package metadata dictionaries
        """
        url = f"{self.base_url}/DeployedPackage/query"

        # Build query filter
        nested_expressions = [
            {
                "argument": ["process"],
                "operator": "EQUALS",
                "property": "componentType",
            },
            {"argument": [True], "operator": "EQUALS", "property": "active"},
        ]

        # Add environment filter if specified
        if environment_name:
            nested_expressions.append(
                {
                    "argument": [environment_name],
                    "operator": "EQUALS",
                    "property": "environmentName",
                }
            )

        query = {
            "QueryFilter": {
                "expression": {
                    "operator": "and",
                    "nestedExpression": nested_expressions,
                }
            }
        }

        try:
            response = requests.post(
                url, headers=self.headers, json=query, auth=self.auth
            )
            response.raise_for_status()

            result = response.json()
            deployed_packages = result.get("result", [])

            print(f"Found {len(deployed_packages)} deployed processes")
            if environment_name:
                print(f"  Filtered to environment: {environment_name}")

            return deployed_packages

        except requests.exceptions.RequestException as e:
            print(f"Error querying deployed packages: {e}")
            if hasattr(e.response, "text"):
                print(f"Response: {e.response.text}")
            return []

    def get_process_name(self, component_id: str) -> str:
        """
        Get the process name from component ID

        Args:
            component_id: The component (process) ID

        Returns:
            Process name
        """
        url = f"{self.base_url}/Process/{component_id}"

        try:
            response = requests.get(url, headers=self.headers, auth=self.auth)
            response.raise_for_status()

            process_data = response.json()
            return process_data.get("name", "Unknown")

        except requests.exceptions.RequestException as e:
            print(f"Error getting process name for {component_id}: {e}")
            return "Unknown"

    def get_package_manifest(self, package_id: str) -> list[dict[str, any]]:
        """
        Get the component manifest for a deployed package

        Args:
            package_id: The package ID from DeployedPackage

        Returns:
            List of component info dictionaries with id and version
        """
        url = f"{self.base_url}/PackagedComponentManifest/{package_id}"

        try:
            response = requests.get(url, headers=self.headers, auth=self.auth)
            response.raise_for_status()

            manifest = response.json()
            component_info = manifest.get("componentInfo", [])

            return component_info

        except requests.exceptions.RequestException as e:
            print(f"Error getting package manifest for {package_id}: {e}")
            if hasattr(e, "response") and e.response is not None:
                print(f"Response: {e.response.text}")
            return []

    def make_bulk_component_metadata_request(
        self, chunked_component_ids: list[str]
    ) -> list[dict[str, any]]:
        """
        Get metadata for multiple components in bulk

        Args:
            component_ids: List of component IDs

        Returns:
            List of component metadata dictionaries
        """

        url = f"{self.base_url}/ComponentMetadata/bulk"

        # Build request body
        request_body = {
            "request": [{"id": comp_id} for comp_id in chunked_component_ids],
            "type": "GET",
        }

        try:
            response = requests.post(
                url, headers=self.headers, json=request_body, auth=self.auth
            )
            response.raise_for_status()

            result = response.json()
            return result.get("response", [])

        except requests.exceptions.RequestException as e:
            print(f"Error getting bulk component metadata: {e}")
            if hasattr(e, "response") and e.response is not None:
                print(f"Response: {e.response.text}")
            return []

    def get_component_metadata_bulk(
        self, component_ids: list[str]
    ) -> list[dict[str, any]]:
        """
        Get metadata for multiple components in bulk

        Args:
            component_ids: List of component IDs

        Returns:
            List of component metadata dictionaries
        """

        chunked_component_ids = chunk_list(component_ids, 100)

        all_metadata = []

        for chunk in chunked_component_ids:
            chunked_component_metadata = self.make_bulk_component_metadata_request(
                chunked_component_ids=chunk
            )
            all_metadata.extend(chunked_component_metadata)

        return all_metadata

    def parse_component_metadata(
        self, metadata: dict[str, any], process_name: str
    ) -> list[dict[str, any]]:
        """
        Parse component metadata to extract connector and operation information

        Args:
            metadata: Component metadata from bulk API
            process_name: Name of the process

        Returns:
            List of dictionaries containing connector/operation metadata
        """
        results = []
        metadata = metadata.get("Result", {})
        try:

            # Get the component type
            component_type = metadata.get("type", "Unknown")
            component_name = metadata.get("name", "Unknown")
            component_id = metadata.get("componentId", "Unknown")
            component_version = metadata.get("version", "Unknown")
            component_current_version = metadata.get("currentVersion", "Unknown")
            component_branch = metadata.get("branchName", "Unknown")

            # For connection components
            if component_type in ["connector-settings", "connector-action"]:
                component_subtype = metadata.get("subType", "Unknown")
                results.append(
                    {
                        "process_name": process_name,
                        "component_id": component_id,
                        "component_name": component_name,
                        "component_type": component_type,
                        "component_subtype": component_subtype,
                        "component_version": component_version,
                        "component_current_version": component_current_version,
                        "component_branch": component_branch,
                    }
                )

        except Exception as e:
            print(f"Error parsing component metadata: {e}")

        return results

    def export_to_csv(self, data: list[dict[str, any]], filename: str = None):
        """
        Export metadata to CSV file

        Args:
            data: List of metadata dictionaries
            filename: Output filename (default: boomi_metadata_TIMESTAMP.csv)
        """
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"boomi_connector_metadata_{timestamp}.csv"

        if not data:
            print("No data to export")
            return

        # Get all unique keys
        fieldnames = list(data[0].keys())
        """
        fieldnames = [
            "process_name",
            "component_name",
            "component_subtype",
            "component_current_version",
            "component_version",
            "component_id",
        ]
        """
        try:
            with open(filename, "w", newline="", encoding="utf-8") as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(data)

            print(f"Data exported to {filename}")

        except IOError as e:
            print(f"Error writing CSV file: {e}")

    def export_to_excel1(self, data: list[dict[str, any]], filename: str = None):
        """
        Export metadata to Excel file with separate sheets for connector-action and connector-settings

        Args:
            data: List of metadata dictionaries
            filename: Output filename (default: boomi_metadata_TIMESTAMP.xlsx)
        """
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"boomi_connector_metadata_{timestamp}.xlsx"

        if not data:
            print("No data to export")
            return

        # Split data by component_type
        connector_actions = [
            row for row in data if row.get("component_type") == "connector-action"
        ]
        connector_settings = [
            row for row in data if row.get("component_type") == "connector-settings"
        ]

        # Define fieldnames
        fieldnames = [
            "process_name",
            "component_name",
            "component_subtype",
            "component_current_version",
            "component_version",
            "component_id",
        ]

        try:
            # Create workbook
            wb = Workbook()

            # Remove default sheet
            wb.remove(wb.active)

            # Create "Operations" sheet for connector-action data
            ws_operations = wb.create_sheet("Operations")
            if connector_actions:
                ws_operations.append(fieldnames)
                for row in connector_actions:
                    ws_operations.append([row.get(field, "") for field in fieldnames])
            else:
                ws_operations.append(fieldnames)  # Write headers even if no data

            # Create "Connections" sheet for connector-settings data
            ws_connections = wb.create_sheet("Connections")
            if connector_settings:
                ws_connections.append(fieldnames)
                for row in connector_settings:
                    ws_connections.append([row.get(field, "") for field in fieldnames])
            else:
                ws_connections.append(fieldnames)  # Write headers even if no data

            # Save workbook
            wb.save(filename)

            print(f"Data exported to {filename}")
            print(f"  - Operations sheet: {len(connector_actions)} rows")
            print(f"  - Connections sheet: {len(connector_settings)} rows")

        except IOError as e:
            print(f"Error writing Excel file: {e}")

    def export_to_excel(self, data: list[dict[str, any]], filename: str = None):
        """
        Export metadata to Excel file with separate sheets for connector-action and connector-settings

        Args:
            data: List of metadata dictionaries
            filename: Output filename (default: boomi_metadata_TIMESTAMP.xlsx)
        """
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"boomi_connector_metadata_{timestamp}.xlsx"

        if not data:
            print("No data to export")
            return

        # Split data by component_type
        connector_actions = [
            row for row in data if row.get("component_type") == "connector-action"
        ]
        connector_settings = [
            row for row in data if row.get("component_type") == "connector-settings"
        ]

        # Define fieldnames
        fieldnames = [
            "process_name",
            "component_name",
            "component_subtype",
            "component_current_version",
            "component_version",
            "component_id",
        ]

        def auto_size_columns(worksheet):
            """Auto-size all columns based on content"""
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)

                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)  # Add padding, cap at 50
                worksheet.column_dimensions[column_letter].width = adjusted_width

        try:
            # Create workbook
            wb = Workbook()

            # Remove default sheet
            wb.remove(wb.active)

            # Create "Operations" sheet for connector-action data
            ws_operations = wb.create_sheet("Operations")
            ws_operations.append(fieldnames)
            for row in connector_actions:
                ws_operations.append([row.get(field, "") for field in fieldnames])

            # Format Operations sheet as table
            if len(connector_actions) > 0:
                table_ref = (
                    f"A1:{chr(65 + len(fieldnames) - 1)}{len(connector_actions) + 1}"
                )
            else:
                table_ref = f"A1:{chr(65 + len(fieldnames) - 1)}1"

            operations_table = Table(displayName="OperationsTable", ref=table_ref)
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            operations_table.tableStyleInfo = style
            ws_operations.add_table(operations_table)

            # Auto-size columns for Operations sheet
            auto_size_columns(ws_operations)

            # Create "Connections" sheet for connector-settings data
            ws_connections = wb.create_sheet("Connections")
            ws_connections.append(fieldnames)
            for row in connector_settings:
                ws_connections.append([row.get(field, "") for field in fieldnames])

            # Format Connections sheet as table
            if len(connector_settings) > 0:
                table_ref = (
                    f"A1:{chr(65 + len(fieldnames) - 1)}{len(connector_settings) + 1}"
                )
            else:
                table_ref = f"A1:{chr(65 + len(fieldnames) - 1)}1"

            connections_table = Table(displayName="ConnectionsTable", ref=table_ref)
            connections_table.tableStyleInfo = style
            ws_connections.add_table(connections_table)

            # Auto-size columns for Connections sheet
            auto_size_columns(ws_connections)

            # Save workbook
            wb.save(filename)

            print(f"Data exported to {filename}")
            print(f"  - Operations sheet: {len(connector_actions)} rows")
            print(f"  - Connections sheet: {len(connector_settings)} rows")

        except IOError as e:
            print(f"Error writing Excel file: {e}")

    def run_extraction(
        self,
        environment_name: str = None,
        export_csv: bool = True,
        csv_filename: str = None,
    ):
        """
        Main method to run the complete extraction process

        Args:
            environment_name: Optional environment name to filter (e.g., "AWS Molecule Prod")
            export_csv: Whether to export results to CSV
            csv_filename: Optional custom CSV filename

        Returns:
            List of all extracted metadata
        """
        print("Starting Boomi metadata extraction...")
        print("=" * 60)

        # Step 1: Query all deployed packages
        print("\nStep 1: Querying deployed packages...")
        deployed_packages = self.query_deployed_packages(environment_name)

        # REMOVE
        deployed_packages = get_sample_deployments_data()

        if not deployed_packages:
            print("No deployed packages found. Exiting.")
            return []

        # Step 2: Download and parse each process
        print(f"\nStep 2: Processing {len(deployed_packages)} deployed packages...")
        all_metadata = []

        for i, package in enumerate(deployed_packages, 1):
            component_id = package.get("componentId")
            package_id = package.get("packageId")
            environment_id = package.get("environmentId")
            deployed_date = package.get("deployedDate", "Unknown")
            deployed_by = package.get("deployedBy", "Unknown")
            # print(package)
            # Get process name
            print(f"  [{i}/{len(deployed_packages)}] Getting process details...")
            process_name = self.get_process_name(component_id)
            print(f"    Process: {process_name}")
            print(f"    Deployed: {deployed_date} by {deployed_by}")

            if not package_id:
                print(f"    Warning: No package ID found for {process_name}")
                continue

            # Get package manifest (all components in this deployment)
            print(f"    Getting package manifest...")
            component_info = self.get_package_manifest(package_id)

            if not component_info:
                print(f"    Warning: No components found in package manifest")
                continue

            print(f"    Found {len(component_info)} components in package")

            # Get all component IDs
            component_ids = [
                comp.get("id") for comp in component_info if comp.get("id")
            ]

            if not component_ids:
                print(f"    Warning: No valid component IDs found")
                continue

            # Get metadata for all components in bulk
            print(f"    Fetching metadata for {len(component_ids)} components...")
            components_metadata = self.get_component_metadata_bulk(component_ids)

            # Parse each component's metadata
            for comp_metadata in components_metadata:
                metadata = self.parse_component_metadata(comp_metadata, process_name)

                # Add deployment information to each record
                for record in metadata:
                    record["deployment_id"] = package.get("deploymentId")
                    record["environment_id"] = environment_id
                    record["deployed_date"] = deployed_date
                    record["deployed_by"] = deployed_by
                    record["package_version"] = package.get("packageVersion")

                all_metadata.extend(metadata)

            print(
                f"    Extracted {len(all_metadata) - len([m for m in all_metadata if m.get('process_name') != process_name])} total items"
            )

        # Step 3: Export results
        print(f"\nStep 3: Processing complete!")
        print(f"Total connector/operations found: {len(all_metadata)}")

        if export_csv and all_metadata:
            print("\nExporting to CSV...")
            # self.export_to_csv(all_metadata, csv_filename)
            self.export_to_excel(
                all_metadata,
                csv_filename.replace(".csv", ".xlsx") if csv_filename else None,
            )
        return all_metadata
